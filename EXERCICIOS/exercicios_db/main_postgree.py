from db_conn_postgree import query_db
import openpyxl

# data_report [0] - Headers
# data_report [1] - data
data_report = query_db("""%PARAM NAME FOR SQL QUERY%""")

ws = openpyxl.Workbook()
ws.create_sheet('Plan1', 0)
plan1 = ws['Plan1']

# Header Report
for column_index, column_names in enumerate(data_report[0], 1):
    plan1.cell(1, column_index).value = column_names

# Inserting data in excel
for row_index, data_row in enumerate(data_report[1], 2):
    for column_index, data_column in enumerate(data_row, 1):
        plan1.cell(row_index, column_index).value = str(data_column)

try:
    ws.save('planilha.xlsx')
    print('Done!')
except Exception as error:
    print(error)
